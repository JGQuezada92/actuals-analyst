"""
Professional Excel Output Generator

Creates investment banking quality Excel workbooks with styled charts.
Matches CFI financial model formatting standards.
"""
import logging
from pathlib import Path
from typing import List, Dict, Any, Optional
from dataclasses import dataclass
from datetime import datetime

logger = logging.getLogger(__name__)

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, NamedStyle
    from openpyxl.chart import BarChart, LineChart, PieChart, Reference
    from openpyxl.chart.series import DataPoint
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.drawing.fill import ColorChoice
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    logger.warning("openpyxl not installed - Excel output disabled")

from config.chart_styles import get_chart_config


@dataclass
class ExcelOutput:
    """Container for Excel output."""
    file_path: str
    sheet_count: int
    chart_count: int


class ExcelGenerator:
    """
    Generate professional Excel workbooks with financial styling.
    """
    
    def __init__(self, output_dir: str = ".outputs"):
        if not HAS_OPENPYXL:
            raise ImportError("Install openpyxl: pip install openpyxl")
        
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self.config = get_chart_config()
        
        # Define reusable styles
        self._setup_styles()
    
    def _setup_styles(self):
        """Create named styles for consistent formatting."""
        # Header style (navy background, white text)
        self.header_fill = PatternFill(
            start_color=self.config.colors.header_bg.lstrip('#'),
            end_color=self.config.colors.header_bg.lstrip('#'),
            fill_type="solid"
        )
        self.header_font = Font(
            name=self.config.typography.family,
            size=self.config.typography.header_size,
            bold=True,
            color="FFFFFF"
        )
        
        # Section header (orange background)
        self.section_fill = PatternFill(
            start_color=self.config.colors.highlight.lstrip('#'),
            end_color=self.config.colors.highlight.lstrip('#'),
            fill_type="solid"
        )
        
        # Data styles
        self.data_font = Font(
            name=self.config.typography.family,
            size=self.config.typography.body_size
        )
        self.data_font_bold = Font(
            name=self.config.typography.family,
            size=self.config.typography.body_size,
            bold=True
        )
        
        # Alternating row fill
        self.alt_row_fill = PatternFill(
            start_color=self.config.colors.alt_row_bg.lstrip('#'),
            end_color=self.config.colors.alt_row_bg.lstrip('#'),
            fill_type="solid"
        )
        
        # Border
        thin_border = Side(style='thin', color='808080')
        self.cell_border = Border(
            bottom=thin_border
        )
    
    def create_financial_report(
        self,
        data: List[Dict[str, Any]],
        title: str,
        columns: List[str] = None,
        include_chart: bool = True,
        chart_type: str = "bar",
        chart_title: str = None,
        value_column: str = "amount",
        category_column: str = None,
    ) -> ExcelOutput:
        """
        Create a professional financial report workbook.
        
        Args:
            data: Data rows to include
            title: Report title
            columns: Columns to include (default: all)
            include_chart: Whether to add a chart
            chart_type: "bar", "line", or "pie"
            chart_title: Chart title (default: same as report title)
            value_column: Column containing values for chart
            category_column: Column for chart categories
        
        Returns:
            ExcelOutput with file path and metadata
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Data"
        
        # Determine columns
        if columns is None and data:
            columns = list(data[0].keys())
        
        # Add title row
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
        title_cell = ws.cell(row=1, column=1, value=title)
        title_cell.font = Font(
            name=self.config.typography.family,
            size=self.config.typography.title_size,
            bold=True,
            color="FFFFFF"
        )
        title_cell.fill = self.header_fill
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 30
        
        # Add subtitle with date
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(columns))
        subtitle = ws.cell(row=2, column=1, value=f"Generated: {datetime.now().strftime('%B %d, %Y')}")
        subtitle.font = Font(
            name=self.config.typography.family,
            size=self.config.typography.small_size,
            color="FFFFFF"
        )
        subtitle.fill = self.header_fill
        subtitle.alignment = Alignment(horizontal='center')
        
        # Add header row
        header_row = 4
        for col_idx, col_name in enumerate(columns, 1):
            cell = ws.cell(row=header_row, column=col_idx, value=col_name)
            cell.font = self.data_font_bold
            cell.border = self.cell_border
            cell.alignment = Alignment(horizontal='center')
        
        # Add data rows
        for row_idx, row_data in enumerate(data, header_row + 1):
            for col_idx, col_name in enumerate(columns, 1):
                value = row_data.get(col_name, "")
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = self.data_font
                
                # Format numbers
                if isinstance(value, (int, float)):
                    if 'amount' in col_name.lower() or 'value' in col_name.lower():
                        cell.number_format = '"$"#,##0_);\\("$"#,##0\\)'
                    elif 'percent' in col_name.lower() or 'rate' in col_name.lower():
                        cell.number_format = '0.0%'
                
                # Alternating row colors
                if row_idx % 2 == 0:
                    cell.fill = self.alt_row_fill
        
        # Auto-fit columns
        for col_idx, col_name in enumerate(columns, 1):
            max_length = max(
                len(str(col_name)),
                max((len(str(row.get(col_name, ""))) for row in data), default=0)
            )
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max_length + 2, 50)
        
        # Add chart if requested
        chart_count = 0
        if include_chart and data:
            chart = self._create_chart(
                ws, 
                data, 
                columns, 
                chart_type, 
                chart_title or title,
                value_column,
                category_column,
                data_start_row=header_row + 1,
                data_end_row=header_row + len(data),
            )
            if chart:
                ws.add_chart(chart, f"A{header_row + len(data) + 3}")
                chart_count = 1
        
        # Save
        safe_title = "".join(c if c.isalnum() else "_" for c in title)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{safe_title}_{timestamp}.xlsx"
        file_path = self.output_dir / filename
        wb.save(file_path)
        
        return ExcelOutput(
            file_path=str(file_path),
            sheet_count=1,
            chart_count=chart_count,
        )
    
    def _create_chart(
        self,
        ws,
        data: List[Dict],
        columns: List[str],
        chart_type: str,
        title: str,
        value_column: str,
        category_column: str,
        data_start_row: int,
        data_end_row: int,
    ):
        """Create a styled chart."""
        # Find column indices
        try:
            value_idx = columns.index(value_column) + 1
        except ValueError:
            # Try to find any amount-like column
            value_idx = next(
                (i + 1 for i, c in enumerate(columns) if 'amount' in c.lower()),
                None
            )
            if value_idx is None:
                return None
        
        cat_idx = None
        if category_column:
            try:
                cat_idx = columns.index(category_column) + 1
            except ValueError:
                pass
        
        if cat_idx is None:
            cat_idx = 1  # Use first column as categories
        
        # Create chart
        if chart_type == "line":
            chart = LineChart()
            chart.style = 10
        elif chart_type == "pie":
            chart = PieChart()
        else:
            chart = BarChart()
            chart.type = "col"
            chart.style = 10
        
        chart.title = title
        chart.width = 15
        chart.height = 7.5
        
        # Set data references
        values = Reference(ws, min_col=value_idx, min_row=data_start_row, max_row=data_end_row)
        categories = Reference(ws, min_col=cat_idx, min_row=data_start_row, max_row=data_end_row)
        
        chart.add_data(values, titles_from_data=False)
        chart.set_categories(categories)
        
        # Style the chart
        if hasattr(chart, 'series') and chart.series:
            series = chart.series[0]
            series.graphicalProperties.solidFill = self.config.colors.primary.lstrip('#')
        
        return chart


def get_excel_generator(output_dir: str = ".outputs") -> Optional[ExcelGenerator]:
    """Get Excel generator instance."""
    if not HAS_OPENPYXL:
        return None
    return ExcelGenerator(output_dir)

